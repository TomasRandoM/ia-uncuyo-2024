from environment import *
from reflexiveAgent import *
from randomAgent import *
from plot import *
#Utilizada para copiar los environment
import copy
#Utilizada para generar el excel
import openpyxl

if __name__ == "__main__":
    #Listas con los tamaños y porcentajes de suciedad
    sizes = [2, 4, 8, 16, 32, 64, 128]
    dirty = [0.1, 0.2, 0.4, 0.8]
    environments = []
    #Se genera una lista de objetos Environment utilizados posteriormente
    for i in sizes:
        for j in dirty:
            for k in range(0, 10):
                environments.append(Environment(i, i, j))
    
    #Se crea el libro excel
    excelBook = openpyxl.Workbook()
    #Se selecciona la hoja activa
    sheet = excelBook.active
    #Se colocan las etiquetas en la primera columna
    sheet.cell(row = 1, column = 1, value = "Tamaño")
    sheet.cell(row = 2, column = 1, value = "TotalSucios")
    sheet.cell(row = 3, column = 1, value = "LimpiadosSimple")
    sheet.cell(row = 4, column = 1, value = "MovimientosSimple")
    sheet.cell(row = 5, column = 1, value = "LimpiadosAleatorio")
    sheet.cell(row = 6, column = 1, value = "MovimientosAleatorio")

    n = len(environments)
    it = 0
    results = []
    resultsRandom = []
    for i in range(2, n + 2):
        #Se realiza una copia del environment para que el mismo entorno sea utilizado por ambos agentes.
        environment1 = copy.deepcopy(environments[i-2])
        #Se guardan los resultados para cada entorno y dirtRate en dos listas con el formato [Entorno, porcentajeSuciedad, celdasLimpiadas, movimientosRealizados]
        if it == 0:
            results.append([str(environment1.sizeX)+ "x" + str(environment1.sizeX), environment1.dirtRate, 0, 0])
            resultsRandom.append([str(environment1.sizeX)+ "x" + str(environment1.sizeX), environment1.dirtRate, 0, 0])
        agent1 = ReflexiveAgent(environments[i-2])
        agent2 = RandomAgent(environment1)
        sheet.cell(row = 1, column = i, value = str(environment1.sizeX) + " x " + str(environment1.sizeY) + " (" + str(environment1.dirtRate) + ")")
        sheet.cell(row = 2, column = i, value = environment1.dirt)
        cleaned, moved = agent1.think()
        #Se suman las celdas limpiadas y movimientos realizados
        results[len(results) - 1][2] += cleaned
        results[len(results) - 1][3] += moved
        sheet.cell(row = 3, column = i, value = cleaned)
        sheet.cell(row = 4, column = i, value = moved)
        cleaned, moved = agent2.think()
        #Se suman las celdas limpiadas y movimientos realizados
        resultsRandom[len(results) - 1][2] += cleaned
        resultsRandom[len(results) - 1][3] += moved
        sheet.cell(row = 5, column = i, value = cleaned)
        sheet.cell(row = 6, column = i, value = moved)
        it += 1
        #Cuando se llega a las 10 iteraciones, se calculan los promedios de los resultados del entorno y dirtRate actual
        if it == 10:
            results[len(results) - 1][2] /= 10
            results[len(results) - 1][3] /= 10
            resultsRandom[len(results) - 1][2] /= 10
            resultsRandom[len(results) - 1][3] /= 10
            it = 0
    #Se llama a la función plotData para que realice los gráficos para cada dirtRate
    plotData(results, resultsRandom, 0.1)
    plotData(results, resultsRandom, 0.2)
    plotData(results, resultsRandom, 0.4)
    plotData(results, resultsRandom, 0.8)

    #Se guarda el archivo excel.
    excelBook.save("./results.xlsx")
    
 